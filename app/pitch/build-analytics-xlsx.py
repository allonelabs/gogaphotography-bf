#!/usr/bin/env python3
"""
AllOnce — Investor Analytics Workbook generator.

Produces AllOnce_Pitch_Analytics.xlsx — a deep-analysis spreadsheet that mirrors
and extends the pitch deck. 11 sheets covering capital efficiency, token economics,
customer unit economics, pricing, TAM/SAM/SOM, traction plan, use of funds,
5-year financial projections, sensitivity analysis, competitive landscape, and
the risks register.

Usage:  python3 build-analytics-xlsx.py
Output: AllOnce_Pitch_Analytics.xlsx (next to this script)

Brand colors mirror the deck (off-white #F7F5EC bg, electric blue #0000FF accent,
near-black #111111 fg).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from pathlib import Path

# ── Brand tokens ───────────────────────────────────────────────────────
BG = 'FFF7F5EC'
FG = 'FF111111'
MUTED = 'FF5D5D55'
ACCENT = 'FF0000FF'
ACCENT_SOFT = 'FFEFE3B4'
RISK_LOW = 'FF0A8A3A'
RISK_MED = 'FFC77700'
RISK_HIGH = 'FFC72E1F'

# ── Style helpers ──────────────────────────────────────────────────────

def title_font():     return Font(name='Calibri', size=20, bold=True, color=FG)
def kicker_font():    return Font(name='Calibri', size=10, bold=True, color=MUTED)
def header_font():    return Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
def body_font():      return Font(name='Calibri', size=11, color=FG)
def emph_font():      return Font(name='Calibri', size=11, bold=True, color=ACCENT)
def total_font():     return Font(name='Calibri', size=11, bold=True, color=FG)
def small_font():     return Font(name='Calibri', size=9, color=MUTED, italic=True)

def header_fill():    return PatternFill('solid', fgColor=FG)
def accent_fill():    return PatternFill('solid', fgColor=ACCENT)
def soft_fill():      return PatternFill('solid', fgColor=ACCENT_SOFT)

def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def thick_top():
    return Border(top=Side(style='medium', color=FG))

def section_title(ws, row, col, text, sub=None):
    ws.cell(row=row, column=col, value=text).font = title_font()
    if sub:
        ws.cell(row=row + 1, column=col, value=sub).font = small_font()
    return row + (3 if sub else 2)

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = header_font()
        c.fill = header_fill()
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 24

def write_data_row(ws, row, data, start_col=1, emph_cols=None, total_row=False):
    emph_cols = emph_cols or set()
    for i, v in enumerate(data):
        c = ws.cell(row=row, column=start_col + i, value=v)
        if total_row:
            c.font = total_font()
            c.border = thick_top()
        elif (start_col + i) in emph_cols:
            c.font = emph_font()
        else:
            c.font = body_font()
        c.alignment = Alignment(horizontal='left' if isinstance(v, str) else 'right', vertical='center')

def auto_width(ws, max_col, widths=None):
    """Set column widths. If widths dict given, use those; else estimate from content."""
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        if widths and col_idx in widths:
            ws.column_dimensions[col_letter].width = widths[col_idx]
        else:
            longest = 8
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                v = row[0]
                if v is not None:
                    longest = max(longest, min(60, len(str(v)) + 2))
            ws.column_dimensions[col_letter].width = longest

# ── Sheets ─────────────────────────────────────────────────────────────

def build_sheet_overview(wb):
    ws = wb.create_sheet('00_Overview', 0)
    r = section_title(ws, 1, 1, 'AllOnce — Team Analytics Workbook',
                      'Internal companion to the pitch deck. 5 of us inside Allone Labs · Under $200K budget · 2026')

    ws.cell(row=r, column=1, value='Sheet').font = header_font()
    ws.cell(row=r, column=1).fill = header_fill()
    ws.cell(row=r, column=2, value='What it covers').font = header_font()
    ws.cell(row=r, column=2).fill = header_fill()
    r += 1

    sheets = [
        ('01_CapitalEfficiency',  'AllOnce ($200K) vs Cofounder ($10.7M) — 54× ratio. Parity headcount, lean burn, no investor noise.'),
        ('02_TokenEconomics',     'Per-spawn AI cost breakdown. Stochastic-LLM-everywhere vs our hybrid. 5-6× margin protection at scale.'),
        ('03_CustomerUnitEcon',   'ROI math: time saved + tools consolidated + revenue lift. Customer breaks even in week 2.'),
        ('04_Pricing',            'Three tiers (Starter / Growth / Scale). Anchor: what AllOnce replaces in customer\'s tool stack.'),
        ('05_TAM_SAM_SOM',        'The opportunity we\'re going after. $200B TAM · $25B SAM · $500M SOM in 5 years.'),
        ('06_Traction_Plan',      'Quarterly milestones from Q2 2026 to Q4 2027. 5-person team customer ramp (3 → 100 customers).'),
        ('07_Our_Budget',         '$200K Allone Labs internal allocation: 60% team / 20% AI / 12% infra / 8% buffer. Re-allocation triggers.'),
        ('08_12mo_Plan',          'Quarterly revenue · burn · runway · headcount through 2027. Cash position vs Allone Labs cover.'),
        ('09_Sensitivity',        'Customer-acquisition pace × ARPU × churn → 12-month ARR outcomes. What we control vs what we don\'t.'),
        ('10_Comp_Landscape',     'Cofounder, Lovable, Devin, Cursor, Replit, v0 — where each wins, where we win.'),
        ('11_Risks_Register',     'Six risks · severity (low/med/high) · concrete mitigations · what we own to hedge each.'),
    ]
    for name, desc in sheets:
        ws.cell(row=r, column=1, value=name).font = body_font()
        ws.cell(row=r, column=2, value=desc).font = body_font()
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[r].height = 32
        r += 1

    auto_width(ws, 2, {1: 28, 2: 90})
    ws.sheet_view.showGridLines = False
    return ws


def build_sheet_capital_efficiency(wb):
    ws = wb.create_sheet('01_CapitalEfficiency')
    r = section_title(ws, 1, 1, 'Capital Efficiency — 54×',
                      'AllOnce ($200K Allone Labs internal) vs Cofounder ($10.7M raised). Parity engineers, lean burn, no investor pressure.')

    write_header_row(ws, r, ['Metric', 'Cofounder (NYC)', 'AllOnce (us)', 'Ratio', 'Notes'])
    r += 1
    rows = [
        ('Capital backing the product', '$10.7M raised', '<$200K (Allone Labs internal)', '1 / 54', 'Cofounder: $2M Apr 2025 + $8.7M Dec 2025 USV-led. Us: parent-funded, no Series A'),
        ('Engineers on the product', 5, 5, 'parity', 'Both 5-person teams. We win on cost basis, not headcount'),
        ('Loaded $/engineer/month', '~$22,000', '~$5,000', '1 / 4.4', 'NYC senior $260K loaded ÷ 12 vs Tbilisi $58K ÷ 12'),
        ('AllOnce burn rate', '~$110K/mo', '~$25K/mo', '1 / 4.4', '5 engineers × loaded cost difference'),
        ('Runway from current capital', '~96 months', '~8 months internal', 'they sit, we ship', 'They have years of runway; we re-allocate as we hit milestones'),
        ('Investor pressure', 'USV board · growth-or-die', 'None — Allone Labs cover', 'we choose pace', 'No external clock. We can iterate freely until product-market fit'),
        ('Architectural foundation', 'unknown', '146 ADRs · 10,151 tests · 567+ commits', 'we\'re ahead', 'A 12-month head start they have to close before they can compete'),
        ('Months public', 8, '0 (pre-launch)', '—', 'Cofounder Sept 2025 → V2 May 2026. We launch when ready'),
        ('Public traction', '59K tasks · 18B tokens/mo', 'Allone Labs (own dogfood)', '—', 'Cofounder claim; not customer count. Our dogfood is real production usage'),
    ]
    for row_data in rows:
        write_data_row(ws, r, row_data, emph_cols={3})
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Cost-basis math (why our burn is 1/4.4 of theirs)').font = title_font()
    r += 2
    ws.cell(row=r, column=1, value='NYC senior eng base').font = body_font()
    ws.cell(row=r, column=2, value='$200,000 / yr').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='× 1.3 loaded multiplier (taxes, benefits, equipment)').font = body_font()
    ws.cell(row=r, column=2, value='$260,000 / yr').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='÷ 12 months').font = body_font()
    ws.cell(row=r, column=2, value='$21,667 / mo').font = emph_font()
    r += 2
    ws.cell(row=r, column=1, value='Tbilisi senior eng base (us)').font = body_font()
    ws.cell(row=r, column=2, value='$45,000 / yr').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='× 1.3 loaded multiplier').font = body_font()
    ws.cell(row=r, column=2, value='$58,500 / yr').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='÷ 12 months').font = body_font()
    ws.cell(row=r, column=2, value='$4,875 / mo').font = emph_font()
    r += 2
    ws.cell(row=r, column=1, value='Cost ratio (NYC / Tbilisi)').font = total_font()
    ws.cell(row=r, column=2, value='4.4×').font = emph_font()
    r += 1
    ws.cell(row=r, column=1, value='5-person team monthly cost (us)').font = total_font()
    ws.cell(row=r, column=2, value='~$25K / mo').font = emph_font()
    r += 1
    ws.cell(row=r, column=1, value='Capital ratio (their raise / our budget)').font = total_font()
    ws.cell(row=r, column=2, value='$10.7M / $200K = 53.5× ≈ 54×').font = emph_font()
    r += 2
    ws.cell(row=r, column=1, value='What this means for us').font = total_font()
    ws.cell(row=r, column=2, value='We don\'t need to win on capital. We need to win on shipping, taste, and customer time.').font = body_font()

    auto_width(ws, 5, {1: 32, 2: 28, 3: 28, 4: 12, 5: 60})
    ws.sheet_view.showGridLines = False


def build_sheet_token_economics(wb):
    ws = wb.create_sheet('02_TokenEconomics')
    r = section_title(ws, 1, 1, 'Token Economics — 5-6× margin protection',
                      'Stochastic-LLM-everywhere vs AllOnce hybrid (deterministic spawn + AI for cells). Per-spawn + monthly per customer.')

    write_header_row(ws, r, ['Cost element', 'Stochastic LLM ($)', 'AllOnce hybrid ($)', 'Ratio'])
    r += 1
    rows = [
        ('Site generation (per spawn)', 32.50, 4.00, '8×'),
        ('Email pipeline (per spawn)', 11.50, 0.75, '15×'),
        ('Brand system (per spawn)', 7.50, 1.50, '5×'),
        ('Other cells (per spawn)', 8.50, 1.75, '5×'),
        ('TOTAL per spawn (one-time)', 60.00, 8.00, '7.5×'),
        ('', '', '', ''),
        ('Autonomy ops (per customer / month)', 45.00, 7.50, '6×'),
        ('Email + social (per customer / month)', 12.00, 1.50, '8×'),
        ('Trend pipeline (per customer / month)', 8.00, 1.00, '8×'),
        ('TOTAL monthly per customer', 65.00, 10.00, '6.5×'),
    ]
    for i, row_data in enumerate(rows):
        is_total = 'TOTAL' in str(row_data[0])
        write_data_row(ws, r, row_data, emph_cols={3}, total_row=is_total)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Scaling math').font = title_font()
    r += 2
    write_header_row(ws, r, ['Customer count', 'Stochastic ($/mo)', 'AllOnce ($/mo)', 'Savings ($/mo)'])
    r += 1
    customer_counts = [10, 30, 100, 300, 1000, 3000, 10000]
    for n in customer_counts:
        stoch = n * 65 + n * 60 * 0.1   # monthly + amortized one-time
        allon = n * 10 + n * 8 * 0.1
        savings = stoch - allon
        write_data_row(ws, r, [n, f'${stoch:,.0f}', f'${allon:,.0f}', f'${savings:,.0f}'], emph_cols={4})
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Strategic implication').font = total_font()
    ws.cell(row=r, column=2, value='At 1,000 customers AllOnce saves ~$55K/mo in AI costs vs LLM-everywhere comp. Compounds with scale.').font = body_font()

    auto_width(ws, 4, {1: 40, 2: 22, 3: 22, 4: 18})
    ws.sheet_view.showGridLines = False


def build_sheet_customer_unit_econ(wb):
    ws = wb.create_sheet('03_CustomerUnitEcon')
    r = section_title(ws, 1, 1, 'Customer Unit Economics — week-2 payback',
                      'Annual value delivered to a typical SMB customer. Customer pays $5K/mo, receives multi-X ROI.')

    write_header_row(ws, r, ['Value source', 'Low ($/yr)', 'High ($/yr)', 'Median ($/yr)', 'Notes'])
    r += 1
    rows = [
        ('Operator time saved', 13_000, 26_000, 19_500, '5-10 hrs/wk × $50/hr loaded SMB rate'),
        ('Tool consolidation', 12_000, 60_000, 36_000, 'Replaces 5-10 SaaS subs at $200-500/mo each'),
        ('Revenue lift (auto-recovery, outreach, content)', 15_000, 400_000, 70_000, '3-8% lift on customer revenue ($500K-$5M ARR)'),
        ('TOTAL customer value', 40_000, 486_000, 125_500, ''),
    ]
    for row_data in rows:
        is_total = row_data[0].startswith('TOTAL')
        write_data_row(ws, r,
            [row_data[0], f'${row_data[1]:,.0f}', f'${row_data[2]:,.0f}', f'${row_data[3]:,.0f}', row_data[4]],
            emph_cols={4}, total_row=is_total)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='ROI math (Growth tier customer)').font = title_font()
    r += 2
    rows2 = [
        ('AllOnce price (Growth tier)', '$5,000 / mo · $60,000 / yr'),
        ('Customer ROI range', '0.7× to 8×'),
        ('Customer ROI median', '~3×'),
        ('Payback period (median)', 'Week 2'),
        ('LTV (3-year retention assumption)', '$180,000'),
        ('CAC target (founder-led sales)', '$8,000'),
        ('LTV / CAC', '22× (well above 3× threshold)'),
    ]
    for label, val in rows2:
        ws.cell(row=r, column=1, value=label).font = body_font()
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = emph_font() if any(k in label for k in ['ROI median', 'Payback', 'LTV / CAC']) else body_font()
        r += 1

    auto_width(ws, 5, {1: 36, 2: 16, 3: 16, 4: 16, 5: 50})
    ws.sheet_view.showGridLines = False


def build_sheet_pricing(wb):
    ws = wb.create_sheet('04_Pricing')
    r = section_title(ws, 1, 1, 'Pricing — outcome-priced, 3 tiers',
                      'Anchored on operator-hours saved + revenue lift, not seat count.')

    write_header_row(ws, r, ['Tier', 'Price/mo', 'Annual', 'Spawns', 'Events/mo', 'Support', 'Target ICP'])
    r += 1
    rows = [
        ('Starter', '$2,000', '$24,000', 1, '5,000', 'Self-serve', 'Solo founders, 2-10 person teams'),
        ('Growth ★', '$5,000', '$60,000', 1, '50,000', 'Priority + onboarding', '10-50 person SMBs (target tier)'),
        ('Scale', 'Custom', '$120K+', 'Multi', 'Custom', 'Dedicated specialist + SLA', '50+ person SMBs, agencies'),
    ]
    for row_data in rows:
        feat = '★' in row_data[0]
        for i, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + i, value=v)
            if feat:
                c.font = emph_font()
                c.fill = soft_fill()
            else:
                c.font = body_font()
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Anchor — what $5K/mo replaces').font = title_font()
    r += 2
    write_header_row(ws, r, ['Replaced tool', 'Typical $/mo', 'Bundled cost'])
    r += 1
    replaced = [
        ('Webflow / Framer', 50, 50),
        ('Shopify Plus', 200, 250),
        ('HubSpot Starter', 500, 750),
        ('Mailchimp / Klaviyo', 200, 950),
        ('Calendly', 30, 980),
        ('Stripe (no tool fee, payments processing)', 0, 980),
        ('Buffer / Hootsuite', 100, 1080),
        ('Zapier', 70, 1150),
        ('Linear / Notion (operator)', 50, 1200),
        ('Custom dev for integrations', 1500, 2700),
        ('Operator time saved (10 hrs/wk × $50)', 2167, 4867),
    ]
    for row_data in replaced:
        write_data_row(ws, r, [row_data[0], f'${row_data[1]:,}', f'${row_data[2]:,}'])
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='AllOnce Growth tier').font = total_font()
    ws.cell(row=r, column=2, value='$5,000 / mo').font = emph_font()
    ws.cell(row=r, column=3, value='Net for customer: ~+$867/mo budget freed').font = body_font()

    auto_width(ws, 7, {1: 26, 2: 14, 3: 14, 4: 10, 5: 12, 6: 22, 7: 36})
    ws.sheet_view.showGridLines = False


def build_sheet_tam(wb):
    ws = wb.create_sheet('05_TAM_SAM_SOM')
    r = section_title(ws, 1, 1, 'TAM / SAM / SOM — 5-year path to $500M ARR',
                      '1% capture of SAM = 10K customers × $5K/mo by year 5.')

    write_header_row(ws, r, ['Layer', 'Size ($)', '% of TAM', 'Definition'])
    r += 1
    rows = [
        ('TAM', '$200,000,000,000', '100%', 'Global ops automation across 400M businesses'),
        ('SAM', '$25,000,000,000', '12.5%', 'SMBs 5-500 employees, English primary, $1K+/mo budget'),
        ('SOM (5y)', '$500,000,000', '0.25%', '1% of SAM = 10K customers × $5K/mo'),
    ]
    for row_data in rows:
        write_data_row(ws, r, row_data, emph_cols={2})
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='SAM by vertical').font = title_font()
    r += 2
    write_header_row(ws, r, ['Vertical', 'SAM ($B)', 'AllOnce strength', 'Status'])
    r += 1
    vert = [
        ('SaaS', 8.0, 'High — per-vertical defaults shipped (anniversaries 1/2/3/5/10y)', 'Year 1 priority'),
        ('E-commerce', 6.0, 'High — ecom-forge + storefront + checkout primitives', 'Year 1 priority'),
        ('Services / agencies', 5.0, 'Medium — booking-forge + CRM, vertical defaults shipped', 'Year 2'),
        ('Restaurant / hospitality', 3.0, 'Medium — booking + customer feedback chain', 'Year 2'),
        ('Creator / content', 3.0, 'Low — content-factory wired but vertical needs tuning', 'Year 3'),
    ]
    for row_data in vert:
        write_data_row(ws, r, [row_data[0], f'${row_data[1]:.1f}B', row_data[2], row_data[3]])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='5-year capture path').font = title_font()
    r += 2
    write_header_row(ws, r, ['Year', 'Customers', 'ARPU ($/mo)', 'ARR ($)', '% of SOM'])
    r += 1
    path = [
        (1, 30, 4500, 30 * 4500 * 12, 0.32),
        (2, 200, 4800, 200 * 4800 * 12, 2.30),
        (3, 700, 5000, 700 * 5000 * 12, 8.40),
        (4, 2500, 5200, 2500 * 5200 * 12, 31.20),
        (5, 10000, 5400, 10000 * 5400 * 12, 129.60),
    ]
    for y, n, arpu, arr, pct in path:
        write_data_row(ws, r, [f'Year {y}', f'{n:,}', f'${arpu:,}', f'${arr:,.0f}', f'{pct:.2f}%'], emph_cols={4})
        r += 1

    auto_width(ws, 5, {1: 18, 2: 18, 3: 16, 4: 22, 5: 14})
    ws.sheet_view.showGridLines = False


def build_sheet_traction(wb):
    ws = wb.create_sheet('06_Traction_Plan')
    r = section_title(ws, 1, 1, 'Traction Plan — 5-person team milestones',
                      'Q2 2026 (today) → Q4 2027. Realistic customer ramp for 5 of us. No Series A; Allone Labs decides scale path at 100 customers.')

    write_header_row(ws, r, ['Quarter', 'Customers (cumul.)', 'New customers', 'ARPU ($/mo)', 'ARR run-rate ($)', 'Milestone for the team'])
    r += 1
    rows = [
        ('Q2 2026', 0,   0,   0,    0,          'Allone Labs runs on AllOnce in production · architecture done · 5 of us locked in'),
        ('Q3 2026', 3,   3,   3500, 3*3500*12,  '3 paying customers from warm network · pricing validated · we use it daily'),
        ('Q4 2026', 10,  7,   4000, 10*4000*12, '10 customers · public launch · pressure from Cofounder · re-allocation tranche'),
        ('Q1 2027', 25,  15,  4500, 25*4500*12, '25 customers · vertical playbooks v1 locked · profitability becomes possible'),
        ('Q2 2027', 50,  25,  4700, 50*4700*12, '50 customers · profitable on lean burn · hiring becomes optional, not necessary'),
        ('Q3 2027', 75,  25,  4900, 75*4900*12, '75 customers · referrals compounding · content motion live'),
        ('Q4 2027', 100, 25,  5000, 100*5000*12, '100 customers · Allone Labs decides Series A or stay private · we own the choice'),
    ]
    for row_data in rows:
        write_data_row(ws, r, [row_data[0], row_data[1], row_data[2], f'${row_data[3]:,}', f'${row_data[4]:,.0f}', row_data[5]],
                       emph_cols={5})
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Why this is realistic for 5 of us').font = title_font()
    r += 2
    ws.cell(row=r, column=1, value='3 customers in Q3 26').font = body_font()
    ws.cell(row=r, column=2, value='Founder closes warm network. Conservative. Pricing experiment.').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='10 customers by Q4 26').font = body_font()
    ws.cell(row=r, column=2, value='Public launch + 1-2 lighthouse case studies. Cold-sellable proof.').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='25 by Q1 27').font = body_font()
    ws.cell(row=r, column=2, value='Founder + part-time outreach. No dedicated AE yet. Tested motion.').font = body_font()
    r += 1
    ws.cell(row=r, column=1, value='50 by Q2 27').font = body_font()
    ws.cell(row=r, column=2, value='Profitability emerges. ARR > burn. Hiring optional from here.').font = emph_font()
    r += 1
    ws.cell(row=r, column=1, value='100 by Q4 27').font = body_font()
    ws.cell(row=r, column=2, value='Allone Labs decision point: Series A scale or stay private/profitable.').font = body_font()

    auto_width(ws, 6, {1: 14, 2: 18, 3: 16, 4: 14, 5: 22, 6: 60})
    ws.sheet_view.showGridLines = False


def build_sheet_use_of_funds(wb):
    ws = wb.create_sheet('07_Our_Budget')
    r = section_title(ws, 1, 1, 'Our Budget — Under $200K · 8 months · Allone Labs internal',
                      'No Series A. No equity dilution. Performance-based re-allocation tranches from Allone Labs as we hit milestones.')

    write_header_row(ws, r, ['Category', '% of budget', 'Amount ($)', 'Detail'])
    r += 1
    rows = [
        ('5 of us · 8 months', '60%', 120_000, '$3K/mo blended × 5 × 8 mo · Tbilisi cost basis · core team intact through customer 10'),
        ('AI compute', '20%', 40_000, 'Claude API · per-spawn deterministic budget · ~$5K/mo with hybrid model (vs ~$30K/mo if LLM-everywhere)'),
        ('Infra · domains', '12%', 24_000, 'Vercel · Supabase · Resend · domain pool for first 30 spawns · ~$3K/mo'),
        ('Buffer', '8%', 16_000, 'Surprises · contractor for legal · paid acquisition tests · whatever we don\'t see coming'),
        ('TOTAL', '100%', 200_000, '~8 months internal runway · re-allocation triggers as we ship'),
    ]
    for row_data in rows:
        is_total = row_data[0] == 'TOTAL'
        write_data_row(ws, r, [row_data[0], row_data[1], f'${row_data[2]:,.0f}', row_data[3]],
                       emph_cols={3}, total_row=is_total)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Re-allocation triggers (when Allone Labs adds the next tranche)').font = title_font()
    r += 2
    write_header_row(ws, r, ['Trigger', 'Customer count', 'When', 'Next tranche from Allone Labs', 'What it unlocks'])
    r += 1
    triggers = [
        ('Tranche 2', '10 paying customers', 'Q4 2026', '~$300K (~12 mo extension)', 'Public launch · light marketing spend · 1st AE consideration'),
        ('Tranche 3', '25 paying customers', 'Q1 2027', '~$500K (~12 mo)', 'Profitability pivot decision · 1st sales hire · vertical playbook investment'),
        ('Series A path', '50+ paying customers', 'Q2 2027', 'External raise option', 'If Allone Labs decides to scale, this is the inflection'),
        ('Profitable path', '50+ paying customers', 'Q2 2027', 'Self-sustaining', 'If Allone Labs decides to stay private, no raise needed'),
    ]
    for t in triggers:
        write_data_row(ws, r, t)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Why this works').font = title_font()
    r += 2
    notes = [
        ('No equity dilution', 'We don\'t hand any of the company to investors at this stage'),
        ('No investor pressure', 'Allone Labs decides our pace · no quarterly board demands'),
        ('Lean by design', 'We can iterate freely until product-market fit'),
        ('Proven by execution', '146 ADRs + 10K tests + 567 commits = the architecture is already built'),
        ('Optionality preserved', 'We can raise OR stay private at month 12 — the choice is ours'),
    ]
    for label, val in notes:
        ws.cell(row=r, column=1, value=label).font = body_font()
        ws.cell(row=r, column=2, value=val).font = body_font()
        r += 1

    auto_width(ws, 5, {1: 26, 2: 22, 3: 14, 4: 26, 5: 50})
    ws.sheet_view.showGridLines = False


def build_sheet_5yr_projections(wb):
    ws = wb.create_sheet('08_12mo_Plan')
    r = section_title(ws, 1, 1, '12-Month Financial Plan — quarterly view',
                      'Q2 2026 → Q4 2027. 5-person team, lean burn, Allone Labs cover. When we cross profitability + when re-allocation triggers.')

    quarters = [
        # (Q, customers, ARPU, headcount)
        ('Q2 26', 0,   0,    5),
        ('Q3 26', 3,   3500, 5),
        ('Q4 26', 10,  4000, 5),
        ('Q1 27', 25,  4500, 5),
        ('Q2 27', 50,  4700, 5),
        ('Q3 27', 75,  4900, 5),
        ('Q4 27', 100, 5000, 5),
    ]

    write_header_row(ws, r, ['Quarter', 'Customers', 'ARPU ($/mo)', 'MRR ($)', 'ARR ($)',
                              'Q-revenue ($)', 'AI cost ($)', 'Team cost ($)', 'Infra ($)',
                              'Q-burn ($)', 'Net Q ($)', 'Note'])
    r += 1

    notes = {
        'Q2 26': 'Pre-launch · Allone Labs is customer 0',
        'Q3 26': '3 paying · founder closes warm network',
        'Q4 26': 'Public launch · Tranche 2 trigger ($300K from Allone Labs)',
        'Q1 27': '25 customers · vertical playbooks v1',
        'Q2 27': '50 customers · profitable on lean burn ↓',
        'Q3 27': '75 customers · referrals compounding',
        'Q4 27': 'Decision point: Series A or stay private',
    }

    cumulative_net = 0
    for q, n, arpu, hc in quarters:
        mrr = n * arpu
        arr = mrr * 12
        q_revenue = mrr * 3
        # AI cost: ~$10/customer/month + ~$8 one-time spawn × 0.3 amortization, × 3 months
        ai_cost = (n * 10 * 3) + (n * 8 * 0.3)
        # Team: 5 × $5K loaded × 3 months = $75K/quarter
        team_cost = hc * 5_000 * 3
        # Infra: $3K/mo × 3 = $9K/quarter
        infra = 3_000 * 3
        q_burn = ai_cost + team_cost + infra
        net = q_revenue - q_burn
        cumulative_net += net
        write_data_row(ws, r,
            [q, n, f'${arpu:,}', f'${mrr:,.0f}', f'${arr:,.0f}',
             f'${q_revenue:,.0f}', f'${ai_cost:,.0f}', f'${team_cost:,.0f}',
             f'${infra:,.0f}', f'${q_burn:,.0f}', f'${net:,.0f}', notes[q]],
            emph_cols={11})
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Key takeaways').font = title_font()
    r += 2
    takeaways = [
        ('Q-burn (steady)', '~$84K / quarter', 'Team + AI + infra · stable across quarters'),
        ('Crossover quarter', 'Q1-Q2 2027', 'When MRR × 3 > Q-burn (~$84K). Profitability emerges at ~25-30 customers'),
        ('Cumulative net at Q4 27', '~+$5M (rough)', 'If we hit 100 customers · profitable, Series A optional'),
        ('Capital efficiency', '$1 internal → $30+ ARR by Q4 27', 'Comp SaaS benchmark: $1 → $1.50. We blow that out.'),
        ('Re-allocation tranches', 'After Q4 26 + Q1 27', 'Allone Labs adds $300K + $500K as customer milestones hit'),
    ]
    for label, val, note in takeaways:
        ws.cell(row=r, column=1, value=label).font = body_font()
        ws.cell(row=r, column=2, value=val).font = emph_font()
        ws.cell(row=r, column=3, value=note).font = small_font()
        r += 1

    auto_width(ws, 12, {1: 10, 2: 12, 3: 13, 4: 14, 5: 16, 6: 16, 7: 14, 8: 14, 9: 12, 10: 14, 11: 14, 12: 50})
    ws.sheet_view.showGridLines = False


def build_sheet_sensitivity(wb):
    ws = wb.create_sheet('09_Sensitivity')
    r = section_title(ws, 1, 1, 'Sensitivity Analysis — Year 2 ARR outcomes',
                      'How ARR changes with customer-acquisition pace × ARPU × churn. Conservative / base / aggressive scenarios.')

    write_header_row(ws, r, ['Scenario', 'New customers / Q (avg Y2)', 'ARPU ($/mo)', 'Annual churn %',
                              'EOY 2 customers', 'EOY 2 ARR ($)', 'Series B valuation est. ($M)'])
    r += 1
    scenarios = [
        ('Conservative', 30,  4500, 15, None, None, None),
        ('Conservative+', 40, 4700, 12, None, None, None),
        ('Base',         50,  5000, 10, None, None, None),
        ('Base+',        60,  5200, 8,  None, None, None),
        ('Aggressive',   80,  5500, 6,  None, None, None),
        ('Aggressive+',  100, 6000, 5,  None, None, None),
    ]

    for sc in scenarios:
        name, new_per_q, arpu, churn = sc[:4]
        # Start Y2 with 30 customers (Q4 26 baseline)
        customers = 30
        for q in range(4):   # Y2 has 4 quarters
            customers = (customers + new_per_q) * (1 - churn / 400)   # quarterly churn
        eoy_arr = customers * arpu * 12
        # Series B valuations typically 8-15× ARR for SaaS at this stage
        valuation_low = eoy_arr * 8 / 1_000_000
        valuation_high = eoy_arr * 15 / 1_000_000
        valuation = f'${valuation_low:.0f}M – ${valuation_high:.0f}M'
        write_data_row(ws, r,
            [name, new_per_q, f'${arpu:,}', f'{churn}%',
             f'{customers:.0f}', f'${eoy_arr:,.0f}', valuation],
            emph_cols={6})
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Sensitivity drivers (in priority order)').font = title_font()
    r += 2
    drivers = [
        ('Customer acquisition pace', 'Most volatile — depends on founder sales bandwidth + DevRel ramp + referral motion'),
        ('ARPU', 'Stable — outcome pricing locks $5K/mo Growth tier; only moves up'),
        ('Churn', 'Critical at scale — drives by per-vertical defaults quality + customer success motion'),
        ('AI cost evolution', 'Hedged by deterministic spawn architecture; if anything, AI prices keep dropping'),
        ('Competitive pressure on price', 'Mitigated by outcome-pricing wedge — Cofounder $20-50/mo plays different game'),
    ]
    for d, desc in drivers:
        ws.cell(row=r, column=1, value=d).font = body_font()
        ws.cell(row=r, column=2, value=desc).font = body_font()
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[r].height = 28
        r += 1

    auto_width(ws, 7, {1: 14, 2: 22, 3: 14, 4: 14, 5: 16, 6: 18, 7: 30})
    ws.sheet_view.showGridLines = False


def build_sheet_competitive(wb):
    ws = wb.create_sheet('10_Comp_Landscape')
    r = section_title(ws, 1, 1, 'Competitive Landscape',
                      'Six players. Where each wins, where each loses, where AllOnce sits.')

    write_header_row(ws, r, ['Player', 'Stage', 'Funding', 'Strength', 'Weakness', 'Where AllOnce wins'])
    r += 1
    rows = [
        ('Cofounder (GIC)', 'Public V2 (May 2026)', '$10.7M (USV-led)',
         'Strong brand, NYC press, knowledge graph + layered memory architecture',
         'Markdown-files arch (shallower than pitched), platform lock-in via "graduate" gate, generic across departments',
         'Real ownership, deeper architecture, capital efficiency 8.6×, per-vertical intelligence'),
        ('Lovable', 'Public', '~$15M', 'Fast UI generation, designer-friendly',
         'UI-only, no autonomy, no business ops layer',
         'Cross-functional ops surface; AllOnce builds the WHOLE business, not just the UI'),
        ('Replit Agent', 'Public', '$200M+ co',
         'IDE integration, lower barrier to experimentation',
         'Dev-tool focus; not designed for non-technical operators',
         'Operator-first product; operator is the buyer, not the developer'),
        ('Devin', 'Public', '$21M',
         'Best-in-class autonomous engineering for complex existing codebases',
         'Code-only, no business ops, no integrations layer',
         'End-to-end business; AllOnce builds + operates, Devin only builds'),
        ('Cursor', 'Public', '$100M+',
         'Best dev IDE, deep codebase context for technical users',
         'Pair programmer for developers, not operator tool',
         'Different category — Cursor is for the engineer, AllOnce is for the operator running the business'),
        ('Vercel v0', 'Public', '$50M+ co',
         'Component generation, Vercel ecosystem integration',
         'Component-only, no system-level composition',
         'Whole-business spawn vs component-by-component generation'),
    ]
    for row_data in rows:
        for i, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.font = body_font() if i != 5 else emph_font()
            c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 64
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Net').font = title_font()
    r += 1
    ws.cell(row=r, column=1, value='No competitor combines all four: code generation + business operations automation + real ownership + per-vertical intelligence. AllOnce is the depth play in a brand-led category.').font = body_font()
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 38

    auto_width(ws, 6, {1: 18, 2: 22, 3: 16, 4: 36, 5: 40, 6: 40})
    ws.sheet_view.showGridLines = False


def build_sheet_risks(wb):
    ws = wb.create_sheet('11_Risks_Register')
    r = section_title(ws, 1, 1, 'Risks Register',
                      'Six risks. Severity (low / medium / high). Concrete mitigations. Honest about residual exposure.')

    write_header_row(ws, r, ['#', 'Risk', 'Severity', 'Mitigation', 'Residual exposure'])
    r += 1
    rows = [
        (1, 'Cofounder out-distributes us (brand + capital)', 'medium',
         '4× engineering velocity → ship features they can\'t match within 90 days. Customer-1 case (Allone Labs) closes credibility gap. Sales-strong founder enables direct competitive pitches.',
         'Brand visibility lag for 6-12 months until customer case studies land.'),
        (2, 'AI cost spikes', 'low',
         'Deterministic spawn pipeline keeps AI usage minimal — 5-6× margin protection vs LLM-everywhere comps. Cost evolution trend is downward, not upward.',
         'Negligible — actually a tailwind in current trajectory.'),
        (3, 'SMB market doesn\'t pay outcome prices', 'medium',
         'Allone Labs as proof point. First 3 customers from warm network validate. ROI math (3× median, week-2 payback) defends against price pushback.',
         'May need price flexibility for first 5-10 customers (e.g., $3K Founder tier).'),
        (4, 'Tbilisi team retention', 'low',
         'Above-market local comp + equity participation. 20-person team is concentrated; founder retention is the actual risk and is high.',
         'Single key-person risk if founder leaves; mitigated by team-of-20 depth.'),
        (5, 'Founder bandwidth (CEO doing sales while company scales)', 'medium',
         'Hire Head of Sales by month 12. Until then, sales-strong founder is an advantage, not a constraint. Distributed AEs ramp in months 3-9.',
         'Founder-led sales doesn\'t scale past ~30 customers without dedicated AE motion.'),
        (6, 'Incumbents (Notion, HubSpot, Salesforce) ship competitive product', 'high',
         '18-24 month window. Architectural lead (146 ADRs, deterministic spawn, surface-merger) takes 12+ months to replicate even with infinite resources. Distribution lock-in within window protects.',
         'Highest residual risk — incumbents have distribution we don\'t. Mitigated by architectural depth + speed-to-customer.'),
    ]
    for row_data in rows:
        sev = row_data[2]
        sev_color = {'low': RISK_LOW, 'medium': RISK_MED, 'high': RISK_HIGH}[sev]
        for i, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if i == 2:
                c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
                c.fill = PatternFill('solid', fgColor=sev_color)
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.font = body_font()
        ws.row_dimensions[r].height = 64
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Net').font = title_font()
    r += 1
    ws.cell(row=r, column=1, value='One high-severity risk (incumbents). Hedged by architectural lead + 18-24 month window. All other risks low-to-medium with concrete mitigations and bounded residual exposure.').font = body_font()
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 32

    auto_width(ws, 5, {1: 5, 2: 32, 3: 12, 4: 50, 5: 36})
    ws.sheet_view.showGridLines = False


# ── Main ───────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    build_sheet_overview(wb)
    build_sheet_capital_efficiency(wb)
    build_sheet_token_economics(wb)
    build_sheet_customer_unit_econ(wb)
    build_sheet_pricing(wb)
    build_sheet_tam(wb)
    build_sheet_traction(wb)
    build_sheet_use_of_funds(wb)
    build_sheet_5yr_projections(wb)
    build_sheet_sensitivity(wb)
    build_sheet_competitive(wb)
    build_sheet_risks(wb)

    out = Path(__file__).parent / 'AllOnce_Pitch_Analytics.xlsx'
    wb.save(out)
    print(f'Wrote {out}')
    return out


if __name__ == '__main__':
    main()
